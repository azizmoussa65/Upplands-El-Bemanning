"""
PIPELINE COMPLET allabolag.se -> Excel enrichi

Fait tout en une seule commande:
  ETAPE 1: scrape allabolag.se (nom, orgnr, telephone, CA, employes, dirigeant, adresse...)
  ETAPE 2: pour chaque entreprise, trouve son site web (API Serper.dev)
  ETAPE 3: sur ce site, extrait l'email du CEO et les numeros de MOBILE (07X)
  ETAPE 4: exporte un seul fichier Excel avec tout

PREREQUIS:
    pip install requests beautifulsoup4 pandas openpyxl
    Cle API gratuite sur https://serper.dev (2500 requetes offertes)

EXEMPLES:
    # 100 entreprises Elinstallationer, toute la Suede
    python pipeline.py --industry-code 10002498 --pages 4 --api-key TA_CLE --output final.xlsx

    # Stockholm uniquement, entreprises de 5 a 50 employes
    python pipeline.py --industry-code 10002498 --county Stockholm --employees-min 5 --employees-max 50 --api-key TA_CLE --output final.xlsx

    # Scraper seulement, sans enrichissement (pas besoin de cle API)
    python pipeline.py --industry-code 10002498 --pages 4 --no-enrich --output brut.xlsx

    # Reprendre un fichier deja enrichi sans regaspiller de credits API
    python pipeline.py --skip-scrape --input final.xlsx --api-key TA_CLE --output final.xlsx --resume
"""

import argparse
import math
import re
import sys
import time
import unicodedata

import pandas as pd
import requests
from bs4 import BeautifulSoup

# =============================================================================
# CONFIGURATION
# =============================================================================

ALLABOLAG_API = "https://www.allabolag.se/api/search"
SERPER_API = "https://google.serper.dev/search"

HEADERS_ALLABOLAG = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/124.0 Safari/537.36",
    "Accept": "application/json",
    "Referer": "https://www.allabolag.se/bransch-sok",
}

HEADERS_BROWSER = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/124.0 Safari/537.36",
}

EMAIL_REGEX = re.compile(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}")

# Mobiles suedois: 07X + 7 chiffres, ou +46 7X... (separateurs libres)
# Exclut volontairement les fixes (08, 019, 031...)
MOBILE_REGEX = re.compile(r"(?:\+46[\s.-]?|0046[\s.-]?|0)7[02369](?:[\s.-]?\d){7}")

# Pages a visiter sur chaque site d'entreprise
CONTACT_PATHS = [
    "", "/kontakt", "/kontakta-oss", "/om-oss", "/about",
    "/contact", "/contact-us", "/medarbetare", "/personal", "/team",
]

# Domaines a ignorer dans les resultats de recherche (annuaires, reseaux sociaux)
IGNORED_DOMAINS = [
    "allabolag.se", "hitta.se", "eniro.se", "facebook.com", "linkedin.com",
    "instagram.com", "merinfo.se", "ratsit.se", "wikipedia.org", "google.com",
    "proff.se", "birinfo.se", "bolagsfakta.se", "kompass.com", "youtube.com",
]

# Prefixes d'emails generiques: ne doivent JAMAIS etre classes comme "email CEO"
GENERIC_PREFIXES = [
    "info", "kontakt", "contact", "kundservice", "support", "order",
    "faktura", "ekonomi", "post", "mail", "hello", "office", "admin",
]

# Colonnes issues de allabolag.se
ALLABOLAG_FIELDS = [
    "name", "legalName", "orgnr", "phone", "phone_looks_valid", "phone2",
    "mobile", "mobile2", "faxNumber", "email", "homePage", "revenue",
    "currency", "profit", "employees", "companyAccountsLastUpdatedDate",
    "contactPerson_name", "contactPerson_role", "contactPerson_birthDate",
    "contactPerson_id", "contactPerson_businessPerson", "currentIndustry_code",
    "currentIndustry_name", "currentIndustry_description", "industries_names",
    "industries_codes", "visitorAddress_line", "visitorAddress_box",
    "visitorAddress_zip", "visitorAddress_city", "postalAddress_line",
    "postalAddress_box", "postalAddress_zip", "postalAddress_city",
    "county", "municipality", "countryPart", "latitude", "longitude",
    "description", "status", "statusRemarks", "certificates", "rating",
    "marketingProtection", "businessUnitType", "companyId", "listingId",
    "customerId",
]

# Colonnes ajoutees par l'enrichissement
ENRICH_FIELDS = [
    "found_website", "found_ceo_email", "found_generic_email",
    "found_mobiles", "best_email", "best_phone", "confidence",
]


# =============================================================================
# OUTILS COMMUNS
# =============================================================================

def clean_phone(phone):
    """Normalise un numero: garde chiffres et + de tete."""
    if not phone:
        return None
    phone = str(phone).strip()
    has_plus = phone.startswith("+")
    digits = re.sub(r"\D", "", phone)
    if not digits:
        return None
    return ("+" if has_plus else "") + digits


def phone_looks_valid(cleaned_phone):
    """Un numero suedois complet a ~10 chiffres. Indicateur pour trier avant appel."""
    if not cleaned_phone:
        return ""
    digits_only = re.sub(r"\D", "", cleaned_phone)
    if cleaned_phone.startswith("+46"):
        return "Oui" if len(digits_only) in (10, 11) else "Non"
    return "Oui" if len(digits_only) == 10 else "Non"


def clean_mobile(raw):
    """Normalise un mobile suedois."""
    raw = str(raw).strip()
    has_plus = raw.startswith("+")
    digits = re.sub(r"\D", "", raw)
    if digits.startswith("0046"):
        digits = "46" + digits[4:]
        has_plus = True
    if has_plus or digits.startswith("46"):
        return "+" + digits
    return digits


def normalize_name(text):
    """Retire accents, minuscules, decoupe en mots (pour comparer nom <-> texte)."""
    if not text:
        return []
    text = unicodedata.normalize("NFKD", str(text)).encode("ascii", "ignore").decode()
    return [w.lower() for w in re.findall(r"[a-zA-Z]+", text) if len(w) > 1]


def is_generic_email(email):
    """True si l'email est du type info@, kontakt@, kundservice@..."""
    if not email:
        return True
    prefix = str(email).split("@")[0].lower()
    return any(prefix.startswith(g) for g in GENERIC_PREFIXES)


def to_number(value):
    """Convertit revenue/employees (parfois '1-4') en float, ou None.
    Un NaN pandas/numpy en entree donnerait sinon un NaN en sortie (str(nan)
    -> 'nan' -> float('nan') ne leve pas d'exception), ce qui corrompt ensuite
    toute somme MongoDB ($sum) et casse le JSON renvoye au frontend."""
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    s = s.split("-")[0].strip().replace(" ", "")
    try:
        result = float(s)
    except ValueError:
        return None
    return None if math.isnan(result) else result


# =============================================================================
# ETAPE 1 : SCRAPING allabolag.se
# =============================================================================

def extract_company_row(company):
    """Aplatit un objet 'company' de l'API allabolag en une ligne (48 colonnes)."""
    contact = company.get("contactPerson") or {}
    current_industry = company.get("currentIndustry") or {}
    location = company.get("location") or {}
    industries = company.get("industries") or []
    visitor = company.get("visitorAddress") or {}
    postal = company.get("postalAddress") or {}
    coords_list = location.get("coordinates") or []
    coords = coords_list[0] if coords_list else {}

    # Adresse principale avec fallback visitor -> postal
    address_line = visitor.get("addressLine") or postal.get("addressLine")
    zip_code = visitor.get("zipCode") or postal.get("zipCode")
    city = visitor.get("postPlace") or postal.get("postPlace")

    status_remarks = company.get("statusRemarks") or []
    certificates = company.get("certificates") or []

    return {
        "name": (company.get("name") or "").strip() or None,
        "legalName": company.get("legalName"),
        "orgnr": company.get("orgnr"),
        "phone": clean_phone(company.get("phone")),
        "phone_looks_valid": phone_looks_valid(clean_phone(company.get("phone"))),
        "phone2": clean_phone(company.get("phone2")),
        "mobile": clean_phone(company.get("mobile")),
        "mobile2": clean_phone(company.get("mobile2")),
        "faxNumber": company.get("faxNumber"),
        "email": company.get("email"),
        "homePage": company.get("homePage"),
        "revenue": company.get("revenue"),
        "currency": company.get("currency"),
        "profit": company.get("profit"),
        "employees": company.get("employees"),
        "companyAccountsLastUpdatedDate": company.get("companyAccountsLastUpdatedDate"),
        "contactPerson_name": contact.get("name"),
        "contactPerson_role": contact.get("role"),
        "contactPerson_birthDate": contact.get("birthDate"),
        "contactPerson_id": contact.get("id"),
        "contactPerson_businessPerson": contact.get("businessPerson"),
        "currentIndustry_code": current_industry.get("code"),
        "currentIndustry_name": current_industry.get("name"),
        "currentIndustry_description": current_industry.get("description"),
        "industries_names": " | ".join(i.get("name") for i in industries if i.get("name")),
        "industries_codes": " | ".join(i.get("code") for i in industries if i.get("code")),
        "visitorAddress_line": address_line,
        "visitorAddress_box": visitor.get("boxAddressLine"),
        "visitorAddress_zip": zip_code,
        "visitorAddress_city": city,
        "postalAddress_line": postal.get("addressLine"),
        "postalAddress_box": postal.get("boxAddressLine"),
        "postalAddress_zip": postal.get("zipCode"),
        "postalAddress_city": postal.get("postPlace"),
        "county": location.get("county"),
        "municipality": location.get("municipality"),
        "countryPart": location.get("countryPart"),
        "latitude": coords.get("ycoordinate"),
        "longitude": coords.get("xcoordinate"),
        "description": company.get("description"),
        "status": company.get("status"),
        "statusRemarks": " | ".join(
            f"{r.get('desc')} ({r.get('date')})" for r in status_remarks if r.get("desc")
        ),
        "certificates": " | ".join(c.get("type") for c in certificates if c.get("type")),
        "rating": company.get("rating"),
        "marketingProtection": company.get("marketingProtection"),
        "businessUnitType": company.get("businessUnitType"),
        "companyId": company.get("companyId"),
        "listingId": company.get("listingId"),
        "customerId": company.get("customerId"),
    }


def fetch_allabolag_page(page, query=None, industry_code=None):
    params = {"page": page}
    if query:
        params["q"] = query
    if industry_code:
        params["industryCode"] = industry_code

    resp = requests.get(ALLABOLAG_API, params=params, headers=HEADERS_ALLABOLAG, timeout=20)
    resp.raise_for_status()
    data = resp.json()

    if isinstance(data, dict):
        return (
            data.get("companies") or data.get("result") or [],
            data.get("pages"),
            data.get("hits"),
        )
    return data, None, None


def scrape_allabolag(args):
    """ETAPE 1: recupere les entreprises depuis allabolag.se, retourne un DataFrame."""
    print("\n" + "=" * 70, file=sys.stderr)
    print("ETAPE 1/3 : Extraction depuis allabolag.se", file=sys.stderr)
    print("=" * 70, file=sys.stderr)

    wanted_counties = None
    if args.county:
        wanted_counties = {c.strip().lower() for c in args.county.split(",")}

    queries = [q.strip() for q in args.query.split(",")] if args.query else [None]
    industry_codes = (
        [c.strip() for c in args.industry_code.split(",")] if args.industry_code else [None]
    )

    all_rows = []
    seen_orgnr = set()

    for q in queries:
        for code in industry_codes:
            label = (f"query={q}" if q else "") + (f" industryCode={code}" if code else "")
            print(f"\n--- Recherche: {label} ---", file=sys.stderr)

            page = args.start_page
            max_pages = args.pages

            while True:
                # Arret si on a atteint la limite d'entreprises demandee
                if args.max_companies and len(all_rows) >= args.max_companies:
                    print(f"Limite de {args.max_companies} entreprises atteinte.", file=sys.stderr)
                    break
                if max_pages is not None and page >= args.start_page + max_pages:
                    break

                print(f"-> Page {page}{f'/{max_pages}' if max_pages else ''} ...", file=sys.stderr)
                try:
                    companies, total_pages, total_hits = fetch_allabolag_page(
                        page, query=q, industry_code=code
                    )
                except requests.RequestException as e:
                    print(f"Erreur page {page}: {e}", file=sys.stderr)
                    break

                if not companies:
                    print("Plus de resultats.", file=sys.stderr)
                    break

                if max_pages is None and total_pages:
                    max_pages = total_pages
                    print(f"   ({total_hits} resultats au total sur {total_pages} pages)", file=sys.stderr)

                for company in companies:
                    if args.max_companies and len(all_rows) >= args.max_companies:
                        break

                    row = extract_company_row(company)
                    orgnr = row.get("orgnr")

                    if orgnr and orgnr in seen_orgnr:
                        continue
                    if wanted_counties and (row.get("county") or "").strip().lower() not in wanted_counties:
                        continue

                    revenue = to_number(row.get("revenue"))
                    if args.revenue_min is not None and (revenue is None or revenue < args.revenue_min):
                        continue
                    if args.revenue_max is not None and (revenue is None or revenue > args.revenue_max):
                        continue

                    employees = to_number(row.get("employees"))
                    if args.employees_min is not None and (employees is None or employees < args.employees_min):
                        continue
                    if args.employees_max is not None and (employees is None or employees > args.employees_max):
                        continue

                    if orgnr:
                        seen_orgnr.add(orgnr)
                    all_rows.append(row)

                page += 1
                time.sleep(args.delay)

    print(f"\n>>> ETAPE 1 terminee: {len(all_rows)} entreprises recuperees.", file=sys.stderr)
    return pd.DataFrame(all_rows, columns=ALLABOLAG_FIELDS)


# =============================================================================
# ETAPE 2 : TROUVER LE SITE WEB (API Serper)
# =============================================================================

def find_website(company_name, api_key, session, debug=False):
    """Cherche le site web officiel via l'API Serper.dev (resultats Google)."""
    try:
        resp = session.post(
            SERPER_API,
            headers={"X-API-KEY": api_key, "Content-Type": "application/json"},
            json={"q": f"{company_name} Sverige", "gl": "se", "hl": "sv"},
            timeout=10,
        )
        resp.raise_for_status()
        data = resp.json()
    except requests.RequestException as e:
        if debug:
            print(f"      [debug] erreur API Serper: {e}", file=sys.stderr)
        return None

    for result in data.get("organic", []):
        link = result.get("link", "")
        if not link:
            continue
        domain = re.sub(r"^https?://(www\.)?", "", link).split("/")[0].lower()
        if any(ignored in domain for ignored in IGNORED_DOMAINS):
            continue
        return link
    return None


# =============================================================================
# ETAPE 3 : EXTRAIRE EMAIL CEO + MOBILES DU SITE
# =============================================================================

def scrape_company_site(base_url, session, manager_name=None):
    """Visite les pages du site et extrait:
       - ceo_email  : email trouve a cote du nom du dirigeant (et non generique)
       - generic_email : info@, kontakt@...
       - mobiles    : liste de numeros 07X trouves
    """
    base_url = base_url.rstrip("/")
    manager_tokens = normalize_name(manager_name) if manager_name else []

    ceo_email = None
    generic_email = None
    mobiles = []

    for path in CONTACT_PATHS:
        url = base_url + path
        try:
            resp = session.get(url, headers=HEADERS_BROWSER, timeout=8)
            if resp.status_code != 200:
                continue
        except requests.RequestException:
            continue

        soup = BeautifulSoup(resp.text, "html.parser")
        page_text = resp.text

        # --- Mobiles: liens tel: + texte brut ---
        for tel in soup.select('a[href^="tel:"]'):
            candidate = tel.get("href", "").replace("tel:", "").strip()
            if MOBILE_REGEX.search(candidate):
                m = clean_mobile(candidate)
                if m and m not in mobiles:
                    mobiles.append(m)
        for match in MOBILE_REGEX.findall(page_text):
            m = clean_mobile(match)
            if m and m not in mobiles:
                mobiles.append(m)

        # --- Email CEO: mailto dont le voisinage contient le nom du dirigeant ---
        if manager_tokens and ceo_email is None:
            for mailto in soup.select('a[href^="mailto:"]'):
                email = mailto.get("href", "").replace("mailto:", "").split("?")[0].strip()
                if not email or is_generic_email(email):
                    continue  # on refuse info@ comme "email CEO"
                context = ""
                parent = mailto.parent
                for _ in range(3):
                    if parent is None:
                        break
                    context += " " + parent.get_text(" ", strip=True)
                    parent = parent.parent
                context_norm = normalize_name(context)
                if any(tok in context_norm for tok in manager_tokens if len(tok) >= 3):
                    ceo_email = email
                    break

        # --- Email generique (fallback) ---
        if generic_email is None:
            mailto = soup.select_one('a[href^="mailto:"]')
            if mailto:
                email = mailto.get("href", "").replace("mailto:", "").split("?")[0].strip()
                if email:
                    generic_email = email
            if generic_email is None:
                match = EMAIL_REGEX.search(page_text)
                if match:
                    candidate = match.group(0)
                    bad_parts = ["example.com", ".png", ".jpg", ".gif", "sentry", "wixpress"]
                    if not any(bad in candidate.lower() for bad in bad_parts):
                        generic_email = candidate

    return {"ceo_email": ceo_email, "generic_email": generic_email, "mobiles": mobiles}


def compute_confidence(company_name, website, ceo_email):
    """Note de fiabilite basique: le domaine correspond-il au nom de l'entreprise ?"""
    if not website:
        return "Aucun site"
    name_tokens = normalize_name(company_name)
    domain = re.sub(r"^https?://(www\.)?", "", str(website)).split("/")[0].split(".")[0].lower()
    matches = any(tok in domain for tok in name_tokens if len(tok) >= 3)
    if matches and ceo_email:
        return "Haute"
    if matches:
        return "Moyenne"
    return "Faible (verifier le site)"


def enrich_dataframe(df, args):
    """ETAPES 2 et 3: enrichit chaque ligne avec site web, email CEO, mobiles."""
    print("\n" + "=" * 70, file=sys.stderr)
    print("ETAPES 2-3/3 : Recherche site web + email CEO + mobiles", file=sys.stderr)
    print("=" * 70, file=sys.stderr)

    for col in ENRICH_FIELDS:
        if col not in df.columns:
            df[col] = None

    session = requests.Session()
    total = len(df)

    try:
        for i, (idx, row) in enumerate(df.iterrows(), start=1):
            name = row.get("name")
            if not name or pd.isna(name):
                continue

            print(f"\n[{i}/{total}] {name}", file=sys.stderr)

            if args.resume and pd.notna(row.get("found_website")):
                print("   -> deja traite, on saute (economise 1 credit API)", file=sys.stderr)
                continue

            try:
                website = find_website(name, args.api_key, session, debug=args.debug)
                df.at[idx, "found_website"] = website

                if not website:
                    print("   -> aucun site trouve", file=sys.stderr)
                    df.at[idx, "confidence"] = "Aucun site"
                else:
                    manager = row.get("contactPerson_name")
                    if pd.isna(manager):
                        manager = None
                    result = scrape_company_site(website, session, manager_name=manager)

                    ceo_email = result["ceo_email"]
                    generic_email = result["generic_email"]
                    mobiles = result["mobiles"]

                    df.at[idx, "found_ceo_email"] = ceo_email
                    df.at[idx, "found_generic_email"] = generic_email
                    df.at[idx, "found_mobiles"] = " | ".join(mobiles) if mobiles else None
                    df.at[idx, "best_email"] = ceo_email or generic_email

                    existing_phone = row.get("phone")
                    if pd.isna(existing_phone):
                        existing_phone = None
                    df.at[idx, "best_phone"] = (mobiles[0] if mobiles else None) or existing_phone
                    df.at[idx, "confidence"] = compute_confidence(name, website, ceo_email)

                    print(f"   site    : {website}", file=sys.stderr)
                    print(f"   CEO     : {ceo_email}", file=sys.stderr)
                    print(f"   general : {generic_email}", file=sys.stderr)
                    print(f"   mobiles : {mobiles if mobiles else 'aucun'}", file=sys.stderr)

            except Exception as row_error:
                print(f"   -> erreur, on continue: {row_error}", file=sys.stderr)

            # Sauvegarde reguliere: le fichier est consultable pendant l'execution
            if i % args.save_every == 0:
                save_excel(df, args.output)
                print(f"   [sauvegarde apres {i}/{total}]", file=sys.stderr)

            time.sleep(args.delay)

    except KeyboardInterrupt:
        print("\n\nInterrompu (Ctrl+C). Sauvegarde de ce qui a ete fait...", file=sys.stderr)
    except Exception as e:
        print(f"\n\nERREUR INATTENDUE: {e}", file=sys.stderr)
        print("Sauvegarde de ce qui a ete fait...", file=sys.stderr)

    return df


# =============================================================================
# EXPORT EXCEL
# =============================================================================

def save_excel(df, path):
    """Sauvegarde en Excel avec les colonnes importantes en premier."""
    priority = [
        "name", "contactPerson_name", "contactPerson_role",
        "best_email", "best_phone", "found_ceo_email", "found_generic_email",
        "found_mobiles", "found_website", "confidence",
        "phone", "employees", "revenue", "county", "municipality",
    ]
    ordered = [c for c in priority if c in df.columns]
    ordered += [c for c in df.columns if c not in ordered]
    df = df[ordered]

    try:
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Entreprises")
            ws = writer.sheets["Entreprises"]
            # En-tetes en gras + largeurs + telephones en format TEXTE (zeros preserves)
            from openpyxl.styles import Font
            for cell in ws[1]:
                cell.font = Font(bold=True)
            wide = {
                "name": 32, "contactPerson_name": 24, "contactPerson_role": 22,
                "best_email": 30, "found_ceo_email": 30, "found_generic_email": 30,
                "found_mobiles": 32, "found_website": 38, "confidence": 22,
                "industries_names": 40,
            }
            phone_cols = {"phone", "phone2", "mobile", "mobile2", "best_phone", "found_mobiles"}
            for i, col in enumerate(df.columns, start=1):
                letter = ws.cell(row=1, column=i).column_letter
                ws.column_dimensions[letter].width = wide.get(col, 16)
                if col in phone_cols:
                    for r in range(2, len(df) + 2):
                        ws[f"{letter}{r}"].number_format = "@"
    except PermissionError:
        print(f"ATTENTION: impossible d'ecrire {path} (fichier ouvert dans Excel ?)", file=sys.stderr)


# =============================================================================
# MAIN
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Pipeline complet: allabolag.se -> site web -> email CEO + mobiles",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )

    # -- Selection des entreprises (etape 1) --
    parser.add_argument("--query", "-q", help="Mot(s)-cle(s), separes par virgule")
    parser.add_argument("--industry-code", help="Code(s) de branche, ex: 10002498 (Elinstallationer)")
    parser.add_argument("--pages", type=int, default=None, help="Nb de pages (25 entreprises/page)")
    parser.add_argument("--start-page", type=int, default=1)
    parser.add_argument("--max-companies", type=int, default=None,
                        help="Nombre MAX d'entreprises a extraire (ex: 100)")
    parser.add_argument("--county", help="Filtrer par region(s), ex: Stockholm")
    parser.add_argument("--revenue-min", type=float)
    parser.add_argument("--revenue-max", type=float)
    parser.add_argument("--employees-min", type=float)
    parser.add_argument("--employees-max", type=float)

    # -- Enrichissement (etapes 2-3) --
    parser.add_argument("--api-key", help="Cle API Serper.dev (requise sauf avec --no-enrich)")
    parser.add_argument("--no-enrich", action="store_true",
                        help="Scraper allabolag seulement, sans chercher les sites/emails")
    parser.add_argument("--resume", action="store_true",
                        help="Saute les lignes ayant deja un found_website (economise des credits)")

    # -- Entree / sortie --
    parser.add_argument("--skip-scrape", action="store_true",
                        help="Sauter l'etape 1 et partir d'un fichier existant (--input)")
    parser.add_argument("--input", help="Fichier .xlsx de depart si --skip-scrape")
    parser.add_argument("--output", default="resultat_final.xlsx")
    parser.add_argument("--delay", type=float, default=1.0, help="Pause entre requetes (s)")
    parser.add_argument("--save-every", type=int, default=10)
    parser.add_argument("--debug", action="store_true")

    args = parser.parse_args()

    # --- Validation des arguments ---
    if args.skip_scrape:
        if not args.input:
            parser.error("--skip-scrape necessite --input fichier.xlsx")
        df = pd.read_excel(args.input)
        print(f"Fichier charge: {len(df)} lignes depuis {args.input}", file=sys.stderr)
        if "name" not in df.columns:
            parser.error("Le fichier d'entree doit contenir une colonne 'name'")
    else:
        if not args.query and not args.industry_code:
            parser.error("Fournir --query ou --industry-code (ou utiliser --skip-scrape --input)")
        df = scrape_allabolag(args)
        if df.empty:
            print("Aucune entreprise recuperee, arret.", file=sys.stderr)
            return
        save_excel(df, args.output)
        print(f"Donnees brutes sauvegardees dans {args.output}", file=sys.stderr)

    # --- Enrichissement ---
    if args.no_enrich:
        print("\n--no-enrich: on s'arrete apres l'extraction allabolag.", file=sys.stderr)
    else:
        if not args.api_key:
            parser.error("--api-key est requise pour l'enrichissement (ou utiliser --no-enrich)")
        df = enrich_dataframe(df, args)

    # --- Sauvegarde finale + resume ---
    save_excel(df, args.output)

    print("\n" + "=" * 70, file=sys.stderr)
    print("TERMINE", file=sys.stderr)
    print("=" * 70, file=sys.stderr)
    print(f"Entreprises          : {len(df)}", file=sys.stderr)
    if "found_website" in df.columns:
        print(f"Sites web trouves    : {df['found_website'].notna().sum()}", file=sys.stderr)
        print(f"Emails (total)       : {df['best_email'].notna().sum()}", file=sys.stderr)
        print(f"  dont emails CEO    : {df['found_ceo_email'].notna().sum()}", file=sys.stderr)
        print(f"Avec mobile(s) 07X   : {df['found_mobiles'].notna().sum()}", file=sys.stderr)
    print(f"\nFichier: {args.output}", file=sys.stderr)


if __name__ == "__main__":
    main()